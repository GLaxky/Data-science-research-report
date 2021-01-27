import datetime
# import numpy as np
import xlrd
import numpy as np
from scipy.stats import norm
import matplotlib.pyplot as plt
from openpyxl import load_workbook

# 显示中文必须的步骤
plt.rcParams['font.sans-serif']=['SimHei']
plt.rcParams['axes.unicode_minus'] = False
sourceFilePath="D:/request/测试文件/weibo.xlsx"

def getData():
    workbook = load_workbook(sourceFilePath)
    # worksheets = workbook.sheets()
    commentsNumList = []
    for sheet in workbook:
        rows = sheet.max_row
        k = range(rows)
        for i in range(1,rows+1):
            print(sheet.cell(row=i, column=4).value)
            commentsNum = sheet.cell(row=i, column=4).value
            if(commentsNum==None):
                continue
            if(str(commentsNum).__len__()==3):
                continue
            if(str(commentsNum).find("100万+")!=-1):
                commentsNumList.append(1000000)
                continue
            commentsNumList.append(int(str(commentsNum)[3:]))
    return commentsNumList

data=getData()
fig, ax = plt.subplots()
ax.hist(data, bins=10000, density=True, alpha=0.6)
mu, std = norm.fit(data)
# Plot the PDF.
# xmin, xmax = plt.xlim()
xmin=0
xmax=1600
x = np.linspace(xmin,xmax , 1000)
p = norm.pdf(x, mu, std)
ax2 = ax.twinx()
# h2, = ax2.plot(x,y1,'r-',linewidth=2,label='y1-中文')
# ax2.set_ylabel('')
ax2.plot(x, p, 'g', linewidth=1,alpha=0.6,label='正态拟合曲线')
ax.set_ylabel('频率/组距')
title = "微博'评论量'直方图: 标准差 = %.2f" % (std)
plt.title(title)
ax.set_xlabel('评论量')
ax.yaxis.label.set_color('b')
ax2.yaxis.label.set_color('g')
plt.axvline(x=352.63,linestyle="--",linewidth=1,color='r',label='均值评论量，352.63')

# my_x_ticks = np.arange(1, 150000, 6000)#原始数据有13个点，故此处为设置从0开始，间隔为1
# plt.xticks(my_x_ticks)
plt.legend(title="注：频率=某种评论量的微博数/有关新冠疫情总微博数")
plt.xlim(0,2500)
plt.show()
